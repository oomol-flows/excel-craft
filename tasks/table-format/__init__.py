#region generated meta
import typing
class Inputs(typing.TypedDict):
    file_path: str
    formatting: dict
    sheet_name: str | None
class Outputs(typing.TypedDict):
    file_path: typing.NotRequired[str]
    success: typing.NotRequired[bool]
    applied_formats: typing.NotRequired[list[str]]
#endregion

from oocana import Context
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os


async def main(params: Inputs, context: Context) -> Outputs:
    """Apply Excel formatting and beautification."""

    file_path = params["file_path"]
    sheet_name = params.get("sheet_name")
    formatting = params["formatting"]

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError("Only Excel files (.xlsx, .xlsm) are supported")

    # Load workbook
    wb = openpyxl.load_workbook(file_path)

    # Select sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    applied_formats = []

    # Apply header style
    if "headerStyle" in formatting:
        _apply_header_style(ws, formatting["headerStyle"])
        applied_formats.append("headerStyle")

    # Apply column widths
    if "columnWidths" in formatting:
        _apply_column_widths(ws, formatting["columnWidths"])
        applied_formats.append("columnWidths")

    # Apply row heights
    if "rowHeights" in formatting:
        _apply_row_heights(ws, formatting["rowHeights"])
        applied_formats.append("rowHeights")

    # Auto-fit columns
    if formatting.get("autoFitColumns", False):
        _auto_fit_columns(ws)
        applied_formats.append("autoFitColumns")

    # Apply conditional formats
    if "conditionalFormats" in formatting:
        _apply_conditional_formats(ws, formatting["conditionalFormats"])
        applied_formats.append("conditionalFormats")

    # Apply freeze panes
    if "freeze" in formatting:
        _apply_freeze(ws, formatting["freeze"])
        applied_formats.append("freeze")

    # Apply data validation
    if "dataValidation" in formatting:
        _apply_data_validation(ws, formatting["dataValidation"])
        applied_formats.append("dataValidation")

    # Save workbook
    wb.save(file_path)

    return {
        "file_path": file_path,
        "success": True,
        "applied_formats": applied_formats
    }


def _apply_header_style(ws, style: dict):
    """Apply style to header row."""

    bold = style.get("bold", True)
    italic = style.get("italic", False)
    font_size = style.get("fontSize", 11)
    font_color = style.get("fontColor", "FFFFFF")
    bg_color = style.get("backgroundColor", "4472C4")
    alignment = style.get("alignment", "center")
    border_style = style.get("borderStyle", "thin")

    # Remove # prefix if present
    if font_color.startswith("#"):
        font_color = font_color[1:]
    if bg_color.startswith("#"):
        bg_color = bg_color[1:]

    # Create font
    font = Font(
        bold=bold,
        italic=italic,
        size=font_size,
        color=font_color
    )

    # Create fill
    fill = PatternFill(
        start_color=bg_color,
        end_color=bg_color,
        fill_type="solid"
    )

    # Create alignment
    align = Alignment(
        horizontal=alignment,
        vertical="center",
        wrap_text=False
    )

    # Create border
    border_map = {
        "thin": "thin",
        "medium": "medium",
        "thick": "thick"
    }
    border_width = border_map.get(border_style, "thin")

    border = Border(
        left=Side(style=border_width),
        right=Side(style=border_width),
        top=Side(style=border_width),
        bottom=Side(style=border_width)
    )

    # Apply to header row (row 1)
    for cell in ws[1]:
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border


def _apply_column_widths(ws, widths: dict):
    """Apply column widths."""

    for col_name, width in widths.items():
        # Find column index by name
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == col_name:
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = width
                break


def _apply_row_heights(ws, heights: dict):
    """Apply row heights."""

    for row_num_str, height in heights.items():
        row_num = int(row_num_str)
        ws.row_dimensions[row_num].height = height


def _auto_fit_columns(ws):
    """Auto-fit column widths based on content."""

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _apply_conditional_formats(ws, formats: list):
    """Apply conditional formatting rules."""

    for fmt in formats:
        range_str = fmt["range"]
        rule_type = fmt["rule"]

        if rule_type == "colorScale":
            colors = fmt.get("colors", ["F8696B", "FFEB84", "63BE7B"])
            rule = ColorScaleRule(
                start_type="min",
                start_color=colors[0],
                mid_type="percentile",
                mid_value=50,
                mid_color=colors[1],
                end_type="max",
                end_color=colors[2]
            )
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "dataBar":
            colors = fmt.get("colors", ["638EC6"])
            rule = DataBarRule(
                start_type="min",
                end_type="max",
                color=colors[0]
            )
            ws.conditional_formatting.add(range_str, rule)

        elif rule_type == "cellValue":
            condition = fmt.get("condition", "> 0")
            style_def = fmt.get("style", {})

            # Parse condition
            operator_map = {
                ">": "greaterThan",
                "<": "lessThan",
                ">=": "greaterThanOrEqual",
                "<=": "lessThanOrEqual",
                "==": "equal",
                "!=": "notEqual"
            }

            for op, op_name in operator_map.items():
                if op in condition:
                    value = condition.split(op)[1].strip()
                    rule = CellIsRule(
                        operator=op_name,
                        formula=[value],
                        fill=PatternFill(start_color=style_def.get("backgroundColor", "FFFF00"), fill_type="solid"),
                        font=Font(color=style_def.get("fontColor", "000000"))
                    )
                    ws.conditional_formatting.add(range_str, rule)
                    break


def _apply_freeze(ws, freeze: dict):
    """Apply freeze panes."""

    row = freeze.get("row", 0)
    col = freeze.get("col", 0)

    # Freeze panes at specified position
    freeze_cell = ws.cell(row + 1, col + 1)
    ws.freeze_panes = freeze_cell


def _apply_data_validation(ws, validations: list):
    """Apply data validation rules."""

    for validation in validations:
        range_str = validation["range"]
        val_type = validation["type"]

        if val_type == "list":
            values = validation.get("values", [])
            formula = f'"{",".join(str(v) for v in values)}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(range_str)

        elif val_type == "whole":
            min_val = validation.get("min")
            max_val = validation.get("max")
            dv = DataValidation(
                type="whole",
                operator="between",
                formula1=min_val,
                formula2=max_val,
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(range_str)

        elif val_type == "decimal":
            min_val = validation.get("min")
            max_val = validation.get("max")
            dv = DataValidation(
                type="decimal",
                operator="between",
                formula1=min_val,
                formula2=max_val,
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(range_str)

        elif val_type == "date":
            min_val = validation.get("min")
            max_val = validation.get("max")
            dv = DataValidation(
                type="date",
                operator="between",
                formula1=min_val,
                formula2=max_val,
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(range_str)
